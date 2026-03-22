#!/usr/bin/env python3
# ============================================================
# generate_mock_data.py
# Sinh dữ liệu mẫu synthetic cho DATN
# 12 tháng 2025: 5 stores, 50 products, 30 customers,
# 20 employees, 10 suppliers, ~10000 invoices
# ============================================================
import os
import sys
import random
import logging
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("mock_data")

# ============================================================
# CONFIG
# ============================================================
random.seed(42)
np.random.seed(42)

OUTPUT_DIR = Path(__file__).parent
OUTPUT_DIR.mkdir(exist_ok=True)

YEAR = 2025
START_DATE = date(YEAR, 1, 1)
END_DATE = date(YEAR, 12, 31)

NUM_STORES = 5
NUM_PRODUCTS = 50
NUM_CUSTOMERS = 30
NUM_EMPLOYEES = 20
NUM_SUPPLIERS = 10

# ============================================================
# MASTER DATA
# ============================================================

STORES = [
    {"MaCH": "CH001", "TenCH": "TechStore Hà Nội", "LoaiHinh": "Flagship", "DiaChi": "123 Trần Duy Hưng, Cầu Giấy", "QuanHuyen": "Cầu Giấy", "ThanhPho": "Hà Nội", "NgayKhaiTruong": "2020-03-15", "QuanLy": "Nguyễn Văn An", "DienTich_m2": 250.0},
    {"MaCH": "CH002", "TenCH": "TechStore Đà Nẵng", "LoaiHinh": "Chi nhánh", "DiaChi": "45 Lê Duẩn, Hải Châu", "QuanHuyen": "Hải Châu", "ThanhPho": "Đà Nẵng", "NgayKhaiTruong": "2021-06-01", "QuanLy": "Trần Thị Bình", "DienTich_m2": 180.0},
    {"MaCH": "CH003", "TenCH": "TechStore TP.HCM", "LoaiHinh": "Flagship", "DiaChi": "78 Nguyễn Trãi, Q.1", "QuanHuyen": "Quận 1", "ThanhPho": "TP.HCM", "NgayKhaiTruong": "2019-01-10", "QuanLy": "Lê Minh Cường", "DienTich_m2": 320.0},
    {"MaCH": "CH004", "TenCH": "TechStore Cần Thơ", "LoaiHinh": "Chi nhánh", "DiaChi": "12 Mậu Thân, Ninh Kiều", "QuanHuyen": "Ninh Kiều", "ThanhPho": "Cần Thơ", "NgayKhaiTruong": "2022-04-20", "QuanLy": "Phạm Thị Dung", "DienTich_m2": 150.0},
    {"MaCH": "CH005", "TenCH": "TechStore Hải Phòng", "LoaiHinh": "Chi nhánh", "DiaChi": "34 Lê Lợi, Lê Chân", "QuanHuyen": "Lê Chân", "ThanhPho": "Hải Phòng", "NgayKhaiTruong": "2023-01-15", "QuanLy": "Hoàng Văn Đức", "DienTich_m2": 140.0},
]

BRANDS = ["Apple", "Samsung", "Dell", "Lenovo", "ASUS", "HP", "Xiaomi", "OPPO", "Vivo", "Realme", "Sony", "LG", "Acer", "Logitech", "JBL", "Anker", "Baseus", "Mophie", "SanDisk", "Western Digital"]
CATEGORIES = ["Điện thoại", "Laptop", "Máy tính bảng", "Tai nghe", "Phụ kiện", "Loa", "Ổ cứng", "USB/TF Card", "Sạc dự phòng", "Camera"]
CATEGORY_SUBCATS = {
    "Điện thoại": ["Smartphone cao cấp", "Smartphone tầm trung", "Smartphone giá rẻ"],
    "Laptop": ["Laptop văn phòng", "Laptop gaming", "Laptop đồ họa", "MacBook"],
    "Máy tính bảng": ["iPad", "Android Tablet"],
    "Tai nghe": ["Tai nghe không dây", "Tai nghe có dây", "Tai nghe gaming"],
    "Phụ kiện": ["Ốp lưng", "Kính cường lực", "Cáp sạc", "Giá đỡ"],
    "Loa": ["Loa Bluetooth", "Loa soundbar", "Loa mini"],
    "Ổ cứng": ["HDD ngoài", "SSD ngoài"],
    "USB/TF Card": ["USB 3.0", "Thẻ nhớ TF"],
    "Sạc dự phòng": ["Pin dự phòng 10000mAh", "Pin dự phòng 20000mAh"],
    "Camera": ["Webcam", "Camera hành trình"],
}

PRODUCT_TEMPLATES = [
    # Điện thoại
    "iPhone 15 Pro Max 256GB", "iPhone 15 Pro 128GB", "iPhone 15 128GB", "iPhone 14 128GB",
    "Samsung Galaxy S24 Ultra", "Samsung Galaxy S24+", "Samsung Galaxy Z Flip5", "Samsung Galaxy A54 5G",
    "Xiaomi 14 Ultra", "Xiaomi Redmi Note 13 Pro", "OPPO Find X7 Ultra", "OPPO Reno11 5G",
    "Vivo X100 Pro", "Realme GT3", "Samsung Galaxy A34 5G",
    # Laptop
    "MacBook Pro 14 M3 Pro", "MacBook Air 15 M3", "MacBook Pro 16 M3 Max",
    "Dell XPS 15 9530", "Dell Inspiron 15 3530", "Dell Latitude 5440",
    "Lenovo ThinkPad X1 Carbon Gen 11", "Lenovo IdeaPad Gaming 3",
    "ASUS ROG Zephyrus G14", "ASUS VivoBook 15 OLED", "ASUS ProArt Studiobook",
    "HP EliteBook 840 G10", "HP Pavilion 15", "HP Omen 16",
    "Acer Swift 3 OLED", "Acer Aspire 5",
    # Tablet
    "iPad Pro 12.9 M4", "iPad Air M2", "iPad mini A17 Pro",
    "Samsung Galaxy Tab S9 Ultra", "Xiaomi Pad 6S Pro",
    # Tai nghe
    "AirPods Pro 2", "AirPods 3", "AirPods 2",
    "Samsung Galaxy Buds2 Pro", "Samsung Galaxy Buds FE",
    "Sony WF-1000XM5", "Sony WH-1000XM5",
    "JBL Tune 770NC", "JBL Flip 6",
    "Xiaomi Buds 4 Pro",
    # Phụ kiện
    "Baseus 3-in-1 Cable", "Anker PowerCore 20000", "Mophie 3-in-1 MagSafe",
    # Loa
    "JBL Charge 5", "JBL PartyBox 110", "Sony SRS-XB13",
    # Ổ cứng
    "WD My Passport 2TB", "Samsung T7 Shield 1TB",
    "SanDisk Extreme Pro 1TB",
]

PAYMENT_METHODS = ["Tiền mặt", "Chuyển khoản", "Quẹt thẻ", "Ví điện tử", "Momo", "VNPay"]
SALES_CHANNELS = ["InStore", "InStore", "InStore", "InStore", "Online"]  # 80% in-store

CITIES = ["Hà Nội", "TP.HCM", "Đà Nẵng", "Cần Thơ", "Hải Phòng", "Hue", "Nha Trang", "Vinh", "Buôn Ma Thuột", "Thanh Hóa"]
CUSTOMER_TYPES = ["Lẻ", "Lẻ", "Lẻ", "Lẻ", "Doanh nghiệp", "VIP"]
GENDERS = ["M", "F", "U"]
SUPPLIER_COUNTRIES = ["Việt Nam", "Trung Quốc", "Hàn Quốc", "Mỹ", "Nhật Bản", "Đài Loan", "Singapore", "Hồng Kông"]
POSITIONS = ["Nhân viên bán hàng", "Nhân viên bán hàng", "Nhân viên bán hàng", "Trưởng ca", "Kỹ thuật viên"]
DEPARTMENTS = ["Kinh doanh", "Kỹ thuật", "Kế toán", "Marketing"]


def generate_products():
    """Sinh danh mục sản phẩm."""
    products = []
    for i in range(NUM_PRODUCTS):
        cat = random.choice(CATEGORIES)
        brand = random.choice(BRANDS)
        subcat = random.choice(CATEGORY_SUBCATS.get(cat, ["Khác"]))

        # Generate name
        template = PRODUCT_TEMPLATES[i] if i < len(PRODUCT_TEMPLATES) else f"{brand} {cat} {i+1}"

        # Price calculation
        if "iPhone" in template or "MacBook" in template or "iPad" in template:
            cost = random.uniform(8_000_000, 45_000_000)
        elif "Samsung Galaxy S" in template or "Samsung Galaxy Z" in template:
            cost = random.uniform(5_000_000, 35_000_000)
        elif "Laptop" in cat or "MacBook" in template:
            cost = random.uniform(10_000_000, 80_000_000)
        elif "Tablet" in cat or "Tab" in template:
            cost = random.uniform(5_000_000, 30_000_000)
        elif "Tai nghe" in cat or "AirPods" in template:
            cost = random.uniform(500_000, 10_000_000)
        else:
            cost = random.uniform(200_000, 8_000_000)

        # List price: cost * margin
        margin = random.uniform(1.15, 1.45)
        list_price = round(cost * margin, -3)

        products.append({
            "MaSP": f"SP{i+1:03d}",
            "TenSP": template,
            "ThuongHieu": brand,
            "DanhMuc": cat,
            "DanhMucCon": subcat,
            "GiaVon": round(cost, -3),
            "GiaNiemYet": list_price,
            "DonViTinh": "cái",
            "BaoHanh_Thang": random.choice([6, 12, 12, 24]),
        })
    return pd.DataFrame(products)


def generate_customers():
    """Sinh danh sách khách hàng."""
    first_names = ["Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Huỳnh", "Vũ", "Đặng", "Bùi", "Đỗ", "Hồ", "Ngô", "Dương", "Lý"]
    mid_names = ["Văn", "Thị", "Minh", "Hoàng", "Thu", "Hữu", "Thanh", "Phương", "Đức", "Lan"]
    last_names = ["An", "Bình", "Cường", "Dung", "Đức", "Hương", "Hùng", "Khoa", "Lan", "Minh", "Nam", "Oanh", "Phương", "Quang", "Sơn", "Thắng", "Trang", "Trung", "Tuấn", "Việt", "Yến", "Anh", "Thơ", "Hà", "Linh", "Mai", "Nga", "Nhi", "Tú", "Vân"]

    customers = []
    for i in range(NUM_CUSTOMERS):
        fn = random.choice(first_names)
        mn = random.choice(mid_names)
        ln = random.choice(last_names)
        fullname = f"{fn} {mn} {ln}"

        join_year = random.randint(2020, YEAR)
        join_month = random.randint(1, 12)
        join_day = random.randint(1, 28)
        join_date = date(join_year, join_month, join_day)

        dob_year = random.randint(1975, 2005)
        dob_month = random.randint(1, 12)
        dob_day = random.randint(1, 28)

        customers.append({
            "MaKH": f"KH{i+1:03d}",
            "HoTen": fullname,
            "GioiTinh": random.choice(GENDERS),
            "NgaySinh": f"{dob_year}-{dob_month:02d}-{dob_day:02d}",
            "DienThoai": f"09{random.randint(0,9)}{random.randint(1000000, 9999999)}",
            "Email": f"kh{i+1}@email.com",
            "ThanhPho": random.choice(CITIES),
            "LoaiKH": random.choice(CUSTOMER_TYPES),
            "DiemTichLuy": random.randint(0, 5000000),
            "NgayDangKy": join_date.strftime("%Y-%m-%d"),
        })
    return pd.DataFrame(customers)


def generate_employees():
    """Sinh danh sách nhân viên."""
    first_names = ["Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Huỳnh", "Vũ", "Đặng", "Bùi", "Đỗ"]
    last_names = ["Anh", "Bình", "Cường", "Dung", "Hương", "Hùng", "Khoa", "Lan", "Minh", "Nam", "Oanh", "Phương", "Sơn", "Thắng", "Trang", "Trung", "Tuấn", "Việt"]

    employees = []
    for i in range(NUM_EMPLOYEES):
        fn = random.choice(first_names)
        ln = random.choice(last_names)
        fullname = f"{fn} {ln}"

        hire_year = random.randint(2019, YEAR)
        hire_month = random.randint(1, 12)
        hire_day = random.randint(1, 28)

        employees.append({
            "MaNV": f"NV{i+1:03d}",
            "HoTen": fullname,
            "PhongBan": random.choice(DEPARTMENTS),
            "ChucVu": random.choice(POSITIONS),
            "MaCH": random.choice(STORES)["MaCH"],
            "NgayVaoLam": date(hire_year, hire_month, hire_day).strftime("%Y-%m-%d"),
        })
    return pd.DataFrame(employees)


def generate_suppliers():
    """Sinh danh sách nhà cung cấp."""
    suppliers = []
    names = [
        "Apple Vietnam", "Samsung Electronics Vietnam", "Dell Technologies Vietnam",
        "Lenovo (Vietnam)", "ASUS Vietnam", "HP Vietnam",
        "Xiaomi Technology Vietnam", "OPPO Vietnam", "Sony Vietnam",
        "Anker Innovations Vietnam",
    ]
    countries = [
        "Mỹ", "Hàn Quốc", "Mỹ", "Hồng Kông", "Đài Loan",
        "Mỹ", "Trung Quốc", "Trung Quốc", "Nhật Bản", "Trung Quốc",
    ]
    for i in range(NUM_SUPPLIERS):
        suppliers.append({
            "MaNCC": f"NCC{i+1:02d}",
            "TenNCC": names[i] if i < len(names) else f"Nhà cung cấp {i+1}",
            "QuocGia": countries[i] if i < len(countries) else random.choice(SUPPLIER_COUNTRIES),
            "NguoiLienHe": f"Mr./Ms. {random.choice(['A', 'B', 'C', 'D'])}",
            "DienThoai": f"0{random.randint(2,9)}{random.randint(10000000, 99999999)}",
            "Email": f"contact{i+1}@supplier.com",
            "DieuKhoanTT_Ngay": random.choice([15, 30, 45, 60]),
        })
    return pd.DataFrame(suppliers)


def generate_sales(customers_df, employees_df, products_df, stores_df):
    """Sinh dữ liệu bán hàng cho cả năm 2025."""
    records = []
    num_days = (END_DATE - START_DATE).days + 1

    # ~10,000 invoices total, spread across 365 days
    # More sales on weekends and near holidays
    daily_avg = 10000 / num_days

    for day_offset in range(num_days):
        current_date = START_DATE + timedelta(days=day_offset)
        dw = current_date.weekday()  # 0=Mon, 6=Sun

        # Seasonal multiplier
        month = current_date.month
        if month in [11, 12]:  # Black Friday, Christmas
            multiplier = 1.5
        elif month in [6, 7, 8]:  # Summer
            multiplier = 1.2
        elif dw >= 5:  # Weekend
            multiplier = 1.4
        else:
            multiplier = 1.0

        # Number of invoices this day
        num_invoices = max(5, int(daily_avg * multiplier + random.gauss(0, daily_avg * 0.3)))
        num_invoices = min(num_invoices, 150)

        for inv_num in range(num_invoices):
            inv_id = f"HD{current_date.strftime('%Y%m%d')}{inv_num+1:04d}"
            store = stores_df.sample(1).iloc[0].to_dict()
            customer = customers_df.sample(1).iloc[0].to_dict()
            emps = employees_df[employees_df["MaCH"] == store["MaCH"]]
            if len(emps) == 0:
                emps = employees_df
            employee = emps.sample(1).iloc[0].to_dict()
            if len(employee) == 0:
                employee = employees_df.sample(1).iloc[0]

            # 1-5 items per invoice
            num_items = random.choices([1, 2, 3, 4, 5], weights=[30, 30, 20, 10, 10])[0]

            for item in range(num_items):
                product = products_df.sample(1).iloc[0]
                qty = random.choices([1, 2, 3, 4, 5], weights=[60, 25, 8, 4, 3])[0]
                unit_price = float(product["GiaNiemYet"])
                discount_pct = random.choices(
                    [0, 0.05, 0.10, 0.15, 0.20],
                    weights=[50, 20, 15, 10, 5]
                )[0]
                discount = unit_price * discount_pct * qty
                tax_pct = 0.10

                records.append({
                    "MaHoaDon": inv_id,
                    "MaSP": product["MaSP"],
                    "MaKH": customer["MaKH"],
                    "MaCH": store["MaCH"],
                    "MaNV": employee["MaNV"],
                    "NgayBan": current_date.strftime("%Y-%m-%d"),
                    "SoLuong": qty,
                    "DonGiaBan": unit_price,
                    "ChietKhau": round(discount, -3),
                    "ThueSuat": tax_pct,
                    "PhuongThucTT": random.choice(PAYMENT_METHODS),
                    "KenhBan": random.choice(SALES_CHANNELS),
                    "IsHoanTra": 0,
                })

    return pd.DataFrame(records)


def generate_inventory(products_df, stores_df):
    """Sinh dữ liệu tồn kho."""
    records = []
    num_days = (END_DATE - START_DATE).days + 1

    for day_offset in range(num_days):
        current_date = START_DATE + timedelta(days=day_offset)
        date_key = int(current_date.strftime("%Y%m%d"))

        for _, store in stores_df.iterrows():
            for _, product in products_df.iterrows():
                # Base stock level depends on product category
                cat = product["DanhMuc"]
                base_stock = {
                    "Điện thoại": 20,
                    "Laptop": 10,
                    "Máy tính bảng": 8,
                    "Tai nghe": 30,
                    "Phụ kiện": 50,
                    "Loa": 15,
                    "Ổ cứng": 25,
                    "USB/TF Card": 40,
                    "Sạc dự phòng": 35,
                    "Camera": 10,
                }.get(cat, 20)

                # Add variation
                closing = max(0, int(base_stock + random.gauss(0, base_stock * 0.3)))

                # Daily movements
                sold = random.randint(0, min(5, closing))
                received = random.randint(0, 10) if random.random() < 0.3 else 0

                opening = max(0, closing + sold - received)
                supplier = random.choice([f"NCC{i+1:02d}" for i in range(NUM_SUPPLIERS)])

                records.append({
                    "MaPhieu": f"PK{date_key}{store['MaCH']}{product['MaSP'][-3:]}",
                    "MaSP": product["MaSP"],
                    "MaCH": store["MaCH"],
                    "MaNCC": supplier,
                    "NgayChot": current_date.strftime("%Y-%m-%d"),
                    "TonDauNgay": opening,
                    "TonCuoiNgay": closing,
                    "NhapTrongNgay": received,
                    "XuatTrongNgay": sold,
                })

    return pd.DataFrame(records)


def save_excel(df, filename, sheet_name):
    """Lưu DataFrame ra Excel với định dạng."""
    filepath = OUTPUT_DIR / filename
    df.to_excel(filepath, sheet_name=sheet_name, index=False, engine="openpyxl")

    # Format
    wb = load_workbook(filepath)
    ws = wb[sheet_name]

    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Auto width
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row=row, column=col)
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    wb.save(filepath)
    logger.info(f"Saved: {filepath} ({len(df)} rows)")


def save_csv(df, filename):
    """Lưu DataFrame ra CSV."""
    filepath = OUTPUT_DIR / filename
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    logger.info(f"Saved: {filepath} ({len(df)} rows)")


def load_workbook(filepath):
    from openpyxl import load_workbook
    return load_workbook(filepath)


def main():
    logger.info("=" * 60)
    logger.info("  MOCK DATA GENERATION STARTED")
    logger.info(f"  Year: {YEAR} | Stores: {NUM_STORES} | Products: {NUM_PRODUCTS}")
    logger.info("=" * 60)

    # Generate all master data
    logger.info("\nGenerating Products...")
    products_df = generate_products()

    logger.info("Generating Customers...")
    customers_df = generate_customers()

    logger.info("Generating Employees...")
    employees_df = generate_employees()

    logger.info("Generating Suppliers...")
    suppliers_df = generate_suppliers()

    # Save dimension data (CSV)
    logger.info("\nSaving dimension data...")
    save_csv(products_df, "DanhMucSanPham.csv")
    save_csv(customers_df, "DanhSachKhachHang.csv")
    save_csv(employees_df, "DanhSachNhanVien.csv")
    save_csv(suppliers_df, "DanhSachNhaCungCap.csv")

    # Save stores (CSV)
    stores_df = pd.DataFrame(STORES)
    save_csv(stores_df, "DanhSachCuaHang.csv")

    # Generate and save transactional data
    logger.info("\nGenerating Sales data (this may take a minute)...")
    sales_df = generate_sales(customers_df, employees_df, products_df, stores_df)
    logger.info(f"  Generated {len(sales_df):,} sales line items")

    # Save sales to Excel (split into monthly sheets)
    logger.info("Saving sales data to Excel...")
    sales_path = OUTPUT_DIR / "BaoCaoDoanhThu_2025.xlsx"

    with pd.ExcelWriter(sales_path, engine="openpyxl") as writer:
        monthly_groups = sales_df.groupby(pd.to_datetime(sales_df["NgayBan"]).dt.to_period("M"))
        for period, group in monthly_groups:
            sheet_name = str(period)[:7]  # e.g. "2025-01"
            group[["MaHoaDon", "MaSP", "MaKH", "MaCH", "MaNV", "NgayBan",
                   "SoLuong", "DonGiaBan", "ChietKhau", "ThueSuat",
                   "PhuongThucTT", "KenhBan", "IsHoanTra"]].to_excel(
                writer, sheet_name=sheet_name, index=False
            )
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

    logger.info(f"Saved: BaoCaoDoanhThu_2025.xlsx ({len(sales_df):,} rows in {len(monthly_groups)} sheets)")

    # Generate and save inventory
    logger.info("\nGenerating Inventory data...")
    inventory_df = generate_inventory(products_df, stores_df)
    logger.info(f"  Generated {len(inventory_df):,} inventory records")

    # Save inventory - monthly
    inv_path = OUTPUT_DIR / "QuanLyKho_2025.xlsx"
    inv_monthly = inventory_df.groupby(pd.to_datetime(inventory_df["NgayChot"]).dt.to_period("M"))
    with pd.ExcelWriter(inv_path, engine="openpyxl") as writer:
        for period, group in inv_monthly:
            sheet_name = str(period)[:7]
            group[["MaPhieu", "MaSP", "MaCH", "MaNCC", "NgayChot",
                   "TonDauNgay", "TonCuoiNgay", "NhapTrongNgay", "XuatTrongNgay"]].to_excel(
                writer, sheet_name=sheet_name, index=False
            )
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color="538135", end_color="538135", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

    logger.info(f"Saved: QuanLyKho_2025.xlsx ({len(inventory_df):,} rows in {len(inv_monthly)} sheets)")

    # Summary
    logger.info("")
    logger.info("=" * 60)
    logger.info("  MOCK DATA GENERATION COMPLETED")
    logger.info("=" * 60)
    logger.info(f"  Products:   {len(products_df):>6,} records")
    logger.info(f"  Customers: {len(customers_df):>6,} records")
    logger.info(f"  Employees: {len(employees_df):>6,} records")
    logger.info(f"  Suppliers: {len(suppliers_df):>6,} records")
    logger.info(f"  Stores:    {len(stores_df):>6,} records")
    logger.info(f"  Sales:     {len(sales_df):>6,} line items")
    logger.info(f"  Inventory: {len(inventory_df):>6,} records")
    logger.info(f"  Output:    {OUTPUT_DIR}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
