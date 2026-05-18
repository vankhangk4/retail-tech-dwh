#!/usr/bin/env python3
"""
Sinh dữ liệu mẫu cho DWH multi-tenant.

Output mặc định: data/data/
- DanhMucSanPham.csv
- DanhMucKhachHang.csv
- DanhMucNhaCungCap.csv
- PhieuNhapHang.csv
- BaoCaoDoanhThu_YYYYMMDD.xlsx
- QuanLyKho_YYYYMMDD.xlsx

Các header được giữ tương thích với ETL hiện tại.
"""

from __future__ import annotations

import csv
import os
import random
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter


OUTPUT_FOLDER = "data"
RANDOM_SEED = 20260518
random.seed(RANDOM_SEED)


@dataclass(frozen=True)
class Product:
    code: str
    name: str
    brand: str
    category: str
    sub_category: str
    cost: int
    list_price: int
    supplier_id: str
    warranty_months: int
    weight: int


@dataclass(frozen=True)
class Customer:
    code: str
    name: str
    gender: str
    dob: date
    city: str
    province: str
    district: str
    phone: str
    email: str
    customer_type: str
    loyalty_points: int
    registered_at: date
    age_group: str
    preferred_channel: str


@dataclass(frozen=True)
class Employee:
    code: str
    name: str
    role: str


SUPPLIERS = [
    ("NCC001", "Công ty TNHH Apple Việt Nam", "Việt Nam", "apple.vn"),
    ("NCC002", "Công ty TNHH Samsung Electronics Việt Nam", "Việt Nam", "samsung.com"),
    ("NCC003", "Công ty TNHH Xiaomi Việt Nam", "Việt Nam", "mi.com"),
    ("NCC004", "Công ty TNHH Digiworld", "Việt Nam", "digiworld.com.vn"),
    ("NCC005", "Công ty Cổ phần Thế Giới Số", "Việt Nam", "dgw.com.vn"),
    ("NCC006", "Công ty TNHH Synnex FPT", "Việt Nam", "synnexfpt.com"),
    ("NCC007", "Công ty TNHH ASUS Việt Nam", "Việt Nam", "asus.com"),
    ("NCC008", "Công ty TNHH Sony Electronics Việt Nam", "Việt Nam", "sony.com.vn"),
    ("NCC009", "Công ty TNHH Logitech Việt Nam", "Thụy Sĩ", "logitech.com"),
    ("NCC010", "Công ty TNHH Anker Innovations", "Trung Quốc", "anker.com"),
]

PRODUCTS = [
    Product("SP001", "iPhone 15 Pro Max 256GB Natural Titanium", "Apple", "Điện thoại", "iPhone", 24450000, 30990000, "NCC001", 12, 42),
    Product("SP002", "iPhone 15 128GB Black", "Apple", "Điện thoại", "iPhone", 15700000, 19990000, "NCC001", 12, 55),
    Product("SP003", "iPhone 14 128GB Blue", "Apple", "Điện thoại", "iPhone", 12800000, 16990000, "NCC001", 12, 45),
    Product("SP004", "Samsung Galaxy S24 Ultra 12GB/256GB", "Samsung", "Điện thoại", "Galaxy S", 22100000, 28990000, "NCC002", 12, 38),
    Product("SP005", "Samsung Galaxy A55 5G 8GB/256GB", "Samsung", "Điện thoại", "Galaxy A", 7850000, 10490000, "NCC002", 12, 72),
    Product("SP006", "Xiaomi Redmi Note 13 Pro 8GB/256GB", "Xiaomi", "Điện thoại", "Redmi Note", 4850000, 7290000, "NCC003", 18, 74),
    Product("SP007", "OPPO Reno11 F 5G 8GB/256GB", "OPPO", "Điện thoại", "Reno", 6200000, 8990000, "NCC004", 12, 46),
    Product("SP008", "MacBook Air 13 M2 8GB/256GB Midnight", "Apple", "Laptop", "MacBook Air", 19800000, 24990000, "NCC001", 12, 35),
    Product("SP009", "MacBook Pro 14 M3 16GB/512GB Space Black", "Apple", "Laptop", "MacBook Pro", 38100000, 48990000, "NCC001", 12, 18),
    Product("SP010", "Dell XPS 13 Plus i7 16GB/512GB", "Dell", "Laptop", "XPS", 26600000, 34990000, "NCC006", 24, 20),
    Product("SP011", "Lenovo ThinkPad X1 Carbon Gen 12", "Lenovo", "Laptop", "ThinkPad", 28700000, 38990000, "NCC006", 24, 16),
    Product("SP012", "ASUS ROG Strix G16 i7 RTX 4060", "ASUS", "Laptop", "Gaming", 30500000, 41990000, "NCC007", 24, 24),
    Product("SP013", "HP Pavilion 15 eg3091TU", "HP", "Laptop", "Pavilion", 13200000, 17990000, "NCC006", 24, 39),
    Product("SP014", "iPad Air 11 M2 Wi-Fi 128GB", "Apple", "Tablet", "iPad Air", 13500000, 16990000, "NCC001", 12, 34),
    Product("SP015", "iPad Pro 11 M4 Wi-Fi 256GB", "Apple", "Tablet", "iPad Pro", 22400000, 28990000, "NCC001", 12, 18),
    Product("SP016", "Samsung Galaxy Tab S9 FE 5G", "Samsung", "Tablet", "Galaxy Tab", 8800000, 11990000, "NCC002", 12, 28),
    Product("SP017", "AirPods Pro 2 USB-C", "Apple", "Tai nghe", "True Wireless", 4350000, 6190000, "NCC001", 12, 86),
    Product("SP018", "Sony WH-1000XM5 Black", "Sony", "Tai nghe", "Chống ồn", 5680000, 7990000, "NCC008", 12, 31),
    Product("SP019", "Samsung Galaxy Buds2 Pro", "Samsung", "Tai nghe", "True Wireless", 2550000, 3990000, "NCC002", 12, 48),
    Product("SP020", "Apple Watch Series 9 GPS 45mm", "Apple", "Đồng hồ", "Apple Watch", 8200000, 10990000, "NCC001", 12, 30),
    Product("SP021", "Samsung Galaxy Watch6 Classic 43mm", "Samsung", "Đồng hồ", "Galaxy Watch", 5100000, 7990000, "NCC002", 12, 28),
    Product("SP022", "Anker PowerCore 10000mAh 22.5W", "Anker", "Phụ kiện", "Sạc dự phòng", 315000, 590000, "NCC010", 18, 95),
    Product("SP023", "Củ sạc Apple USB-C 20W", "Apple", "Phụ kiện", "Củ sạc", 315000, 590000, "NCC001", 12, 90),
    Product("SP024", "Cáp USB-C to USB-C Anker 1.8m", "Anker", "Phụ kiện", "Cáp sạc", 95000, 190000, "NCC010", 18, 100),
    Product("SP025", "Ốp lưng iPhone 15 Pro Max MagSafe", "Apple", "Phụ kiện", "Ốp lưng", 280000, 690000, "NCC001", 12, 80),
    Product("SP026", "Miếng dán cường lực iPhone 15 Pro Max", "Spigen", "Phụ kiện", "Miếng dán", 55000, 150000, "NCC005", 6, 100),
    Product("SP027", "Chuột Logitech MX Master 3S Graphite", "Logitech", "Phụ kiện", "Chuột", 1750000, 2590000, "NCC009", 12, 24),
    Product("SP028", "Bàn phím Logitech MX Keys S", "Logitech", "Phụ kiện", "Bàn phím", 1660000, 2390000, "NCC009", 12, 22),
    Product("SP029", "Màn hình Dell UltraSharp U2723QE 27 inch", "Dell", "Màn hình", "Văn phòng", 9400000, 12990000, "NCC006", 36, 12),
    Product("SP030", "Router Wi-Fi 6 TP-Link Archer AX55", "TP-Link", "Thiết bị mạng", "Router", 1680000, 2490000, "NCC004", 24, 26),
]

EMPLOYEES = [
    Employee("NV001", "Nguyễn Minh Quân", "Cửa hàng trưởng"),
    Employee("NV002", "Trần Thu Hà", "Tư vấn bán hàng"),
    Employee("NV003", "Lê Hoàng Nam", "Tư vấn bán hàng"),
    Employee("NV004", "Phạm Gia Huy", "Thu ngân"),
    Employee("NV005", "Đặng Ngọc Anh", "Tư vấn bán hàng"),
    Employee("NV006", "Vũ Khánh Linh", "Chăm sóc khách hàng"),
    Employee("NV007", "Bùi Đức Anh", "Tư vấn kỹ thuật"),
    Employee("NV008", "Hoàng Phương Thảo", "Tư vấn bán hàng"),
]

CITY_PROFILE = [
    ("Hà Nội", "Hà Nội", ["Cầu Giấy", "Đống Đa", "Ba Đình", "Thanh Xuân", "Hoàng Mai", "Nam Từ Liêm"], 24),
    ("TP. Hồ Chí Minh", "TP. Hồ Chí Minh", ["Quận 1", "Quận 3", "Bình Thạnh", "Tân Bình", "Phú Nhuận", "Thủ Đức"], 26),
    ("Đà Nẵng", "Đà Nẵng", ["Hải Châu", "Thanh Khê", "Sơn Trà", "Cẩm Lệ"], 8),
    ("Hải Phòng", "Hải Phòng", ["Ngô Quyền", "Lê Chân", "Hồng Bàng", "Kiến An"], 6),
    ("Cần Thơ", "Cần Thơ", ["Ninh Kiều", "Cái Răng", "Bình Thủy"], 6),
    ("Bình Dương", "Bình Dương", ["Thủ Dầu Một", "Dĩ An", "Thuận An"], 7),
    ("Đồng Nai", "Đồng Nai", ["Biên Hòa", "Long Khánh", "Trảng Bom"], 6),
    ("Khánh Hòa", "Khánh Hòa", ["Nha Trang", "Cam Ranh", "Diên Khánh"], 5),
    ("Quảng Ninh", "Quảng Ninh", ["Hạ Long", "Cẩm Phả", "Uông Bí"], 4),
    ("Nghệ An", "Nghệ An", ["Vinh", "Cửa Lò", "Diễn Châu"], 4),
    ("Thừa Thiên Huế", "Thừa Thiên Huế", ["Huế", "Hương Thủy", "Phú Vang"], 4),
]

LAST_NAMES = [
    ("Nguyễn", 34), ("Trần", 12), ("Lê", 10), ("Phạm", 7), ("Hoàng", 5),
    ("Huỳnh", 4), ("Phan", 4), ("Vũ", 4), ("Võ", 3), ("Đặng", 3),
    ("Bùi", 3), ("Đỗ", 3), ("Hồ", 2), ("Ngô", 2), ("Dương", 2), ("Lý", 1),
]

MALE_MIDDLE = ["Văn", "Minh", "Quang", "Đức", "Gia", "Hữu", "Tuấn", "Hoàng", "Anh", "Thành", "Nhật", "Khánh"]
FEMALE_MIDDLE = ["Thị", "Thu", "Ngọc", "Phương", "Thanh", "Mai", "Kim", "Bảo", "Khánh", "Thùy", "Hồng", "Gia"]
MALE_GIVEN = ["An", "Bảo", "Cường", "Dũng", "Đạt", "Đức", "Hải", "Hiếu", "Huy", "Khang", "Khôi", "Long", "Minh", "Nam", "Phúc", "Quân", "Sơn", "Thắng", "Tín", "Tuấn", "Việt"]
FEMALE_GIVEN = ["Anh", "Châu", "Chi", "Dung", "Giang", "Hà", "Hạnh", "Hương", "Linh", "Loan", "Mai", "My", "Ngân", "Nhi", "Phương", "Quỳnh", "Thảo", "Trang", "Trâm", "Vy", "Yến"]

EMAIL_DOMAINS = ["gmail.com", "gmail.com", "gmail.com", "outlook.com", "icloud.com", "yahoo.com"]
PAYMENT_METHODS = ["Tiền mặt", "Chuyển khoản", "Thẻ", "Ví điện tử"]
SALES_CHANNELS = ["POS", "Online", "Telesales", "Marketplace"]
CUSTOMER_TYPES = ["Khách lẻ", "Thành viên", "VIP", "Khách doanh nghiệp"]


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    without_marks = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return without_marks.replace("đ", "d").replace("Đ", "D")


def slugify(value: str) -> str:
    value = strip_accents(value).lower()
    value = re.sub(r"[^a-z0-9]+", ".", value)
    return value.strip(".")


def weighted_choice(items):
    values, weights = zip(*items)
    return random.choices(values, weights=weights, k=1)[0]


def random_phone() -> str:
    prefixes = [
        "032", "033", "034", "035", "036", "037", "038", "039",
        "070", "076", "077", "078", "079",
        "081", "082", "083", "084", "085",
        "088", "090", "091", "092", "093", "094", "096", "097", "098",
    ]
    return f"{random.choice(prefixes)}{random.randint(1000000, 9999999)}"


def random_date(start: date, end: date) -> date:
    return start + timedelta(days=random.randint(0, (end - start).days))


def format_date(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def format_datetime(value: datetime) -> str:
    return value.strftime("%d/%m/%Y %H:%M:%S")


def realistic_sale_time(day: date) -> datetime:
    if random.random() < 0.68:
        hour = random.randint(18, 21)
    elif random.random() < 0.55:
        hour = random.randint(10, 12)
    else:
        hour = random.randint(13, 17)
    return datetime.combine(day, time(hour, random.randint(0, 59), random.randint(0, 59)))


def age_group_from_year(year: int) -> str:
    age = 2026 - year
    if age < 25:
        return "18-24"
    if age < 35:
        return "25-34"
    if age < 45:
        return "35-44"
    if age < 55:
        return "45-54"
    return "55+"


def make_person_name(gender: str) -> tuple[str, str, str]:
    last_name = weighted_choice(LAST_NAMES)
    if gender == "Nam":
        middle = random.choice(MALE_MIDDLE)
        given = random.choice(MALE_GIVEN)
    else:
        middle = random.choice(FEMALE_MIDDLE)
        given = random.choice(FEMALE_GIVEN)
    return f"{last_name} {middle} {given}", middle, given


def make_email(full_name: str, phone: str, index: int) -> str:
    parts = slugify(full_name).split(".")
    domain = random.choice(EMAIL_DOMAINS)
    tail = phone[-3:] if random.random() < 0.45 else str(1980 + index % 30)
    pattern = random.choice([
        "{given}.{last}{tail}",
        "{last}.{given}{tail}",
        "{full}{tail}",
        "{given}{tail}",
    ])
    last = parts[0]
    given = parts[-1]
    full = "".join(parts)
    return pattern.format(last=last, given=given, full=full, tail=tail) + f"@{domain}"


def write_csv(path: str, rows: list[dict]) -> None:
    if not rows:
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2EF")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_len + 2, 12), 34)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    ws.freeze_panes = "A2"


def build_customers(num_customers: int = 220) -> list[Customer]:
    customers: list[Customer] = []
    registered_start = date(2023, 1, 1)
    registered_end = date(2026, 1, 15)
    used_emails = set()

    for index in range(1, num_customers + 1):
        gender = weighted_choice([("Nam", 48), ("Nữ", 52)])
        full_name, _, _ = make_person_name(gender)
        city, province, districts = weighted_choice([(c[:3], c[3]) for c in CITY_PROFILE])
        district = random.choice(districts)
        birth_year = random.choices(
            [random.randint(1968, 1978), random.randint(1979, 1989), random.randint(1990, 1999), random.randint(2000, 2007)],
            weights=[12, 24, 42, 22],
            k=1,
        )[0]
        dob = date(birth_year, random.randint(1, 12), random.randint(1, 28))
        phone = random_phone()
        email = make_email(full_name, phone, index)
        while email in used_emails:
            email = make_email(full_name, phone, index + random.randint(10, 999))
        used_emails.add(email)

        customer_type = weighted_choice([
            ("Khách lẻ", 36),
            ("Thành viên", 46),
            ("VIP", 12),
            ("Khách doanh nghiệp", 6),
        ])
        loyalty_multiplier = {
            "Khách lẻ": 1,
            "Thành viên": 4,
            "VIP": 11,
            "Khách doanh nghiệp": 8,
        }[customer_type]
        registered_at = random_date(registered_start, registered_end)

        customers.append(Customer(
            code=f"KH{index:05d}",
            name=full_name,
            gender=gender,
            dob=dob,
            city=city,
            province=province,
            district=district,
            phone=phone,
            email=email,
            customer_type=customer_type,
            loyalty_points=random.randint(80, 1800) * loyalty_multiplier,
            registered_at=registered_at,
            age_group=age_group_from_year(birth_year),
            preferred_channel=weighted_choice([("POS", 58), ("Online", 28), ("Telesales", 7), ("Marketplace", 7)]),
        ))

    return customers


def generate_product_csv(output_dir: str) -> list[dict]:
    rows = []
    for product in PRODUCTS:
        rows.append({
            "Mã SP": product.code,
            "Tên SP": product.name,
            "Thương hiệu": product.brand,
            "Danh mục": product.category,
            "Danh mục con": product.sub_category,
            "Giá vốn": product.cost,
            "Giá niêm yết": product.list_price,
            "Đơn vị tính": "Cái",
            "Bảo hành (tháng)": product.warranty_months,
            "Mã NCC": product.supplier_id,
        })

    csv_path = os.path.join(output_dir, "DanhMucSanPham.csv")
    write_csv(csv_path, rows)
    print(f"  [OK] {csv_path} ({len(rows)} sản phẩm)")
    return rows


def generate_supplier_csv(output_dir: str) -> list[dict]:
    rows = []
    contact_names = [
        "Nguyễn Hải Anh", "Trần Minh Đức", "Lê Thùy Dương", "Phạm Quang Huy",
        "Hoàng Khánh Linh", "Bùi Gia Bảo", "Vũ Ngọc Mai", "Đặng Quốc Việt",
    ]
    for index, (code, name, country, domain) in enumerate(SUPPLIERS, start=1):
        contact = contact_names[(index - 1) % len(contact_names)]
        rows.append({
            "Mã NCC": code,
            "Tên NCC": name,
            "Quốc gia": country,
            "Điều khoản TT (ngày)": random.choice([15, 30, 30, 45, 60]),
            "Số điện thoại": random_phone(),
            "Email": f"{slugify(contact).replace('.', '')}@{domain}",
            "Người liên hệ": contact,
            "Thành phố": random.choice(["TP. Hồ Chí Minh", "Hà Nội", "Đà Nẵng", "Bình Dương"]),
        })

    csv_path = os.path.join(output_dir, "DanhMucNhaCungCap.csv")
    write_csv(csv_path, rows)
    print(f"  [OK] {csv_path} ({len(rows)} nhà cung cấp)")
    return rows


def generate_customer_csv(output_dir: str, num_customers: int = 220) -> list[Customer]:
    customers = build_customers(num_customers)
    rows = []
    for customer in customers:
        rows.append({
            "Mã KH": customer.code,
            "Họ tên": customer.name,
            "Giới tính": customer.gender,
            "Ngày sinh": format_date(customer.dob),
            "Thành phố": customer.city,
            "Tỉnh/TP": customer.province,
            "Quận/Huyện": customer.district,
            "Số điện thoại": customer.phone,
            "Email": customer.email,
            "Loại KH": customer.customer_type,
            "Điểm tích lũy": customer.loyalty_points,
            "Ngày đăng ký": format_date(customer.registered_at),
            "Nhóm tuổi": customer.age_group,
            "Kênh ưu tiên": customer.preferred_channel,
        })

    csv_path = os.path.join(output_dir, "DanhMucKhachHang.csv")
    write_csv(csv_path, rows)
    print(f"  [OK] {csv_path} ({len(rows)} khách hàng)")
    return customers


def choose_customer(customers: list[Customer], channel: str) -> Customer | None:
    if random.random() < 0.1:
        return None

    type_weights = {
        "POS": {"Khách lẻ": 45, "Thành viên": 42, "VIP": 8, "Khách doanh nghiệp": 5},
        "Online": {"Khách lẻ": 28, "Thành viên": 48, "VIP": 16, "Khách doanh nghiệp": 8},
        "Telesales": {"Khách lẻ": 10, "Thành viên": 36, "VIP": 24, "Khách doanh nghiệp": 30},
        "Marketplace": {"Khách lẻ": 56, "Thành viên": 34, "VIP": 6, "Khách doanh nghiệp": 4},
    }
    target_type = weighted_choice(list(type_weights.get(channel, type_weights["POS"]).items()))
    pool = [customer for customer in customers if customer.customer_type == target_type]
    return random.choice(pool or customers)


def choose_products_for_invoice(channel: str) -> list[Product]:
    category_weights = {
        "POS": [("Điện thoại", 42), ("Laptop", 16), ("Tablet", 8), ("Tai nghe", 11), ("Đồng hồ", 8), ("Phụ kiện", 14), ("Màn hình", 1), ("Thiết bị mạng", 0)],
        "Online": [("Điện thoại", 31), ("Laptop", 15), ("Tablet", 9), ("Tai nghe", 13), ("Đồng hồ", 8), ("Phụ kiện", 20), ("Màn hình", 2), ("Thiết bị mạng", 2)],
        "Telesales": [("Điện thoại", 28), ("Laptop", 26), ("Tablet", 10), ("Tai nghe", 6), ("Đồng hồ", 5), ("Phụ kiện", 10), ("Màn hình", 9), ("Thiết bị mạng", 6)],
        "Marketplace": [("Điện thoại", 25), ("Laptop", 5), ("Tablet", 4), ("Tai nghe", 18), ("Đồng hồ", 5), ("Phụ kiện", 40), ("Màn hình", 1), ("Thiết bị mạng", 2)],
    }
    category = weighted_choice(category_weights.get(channel, category_weights["POS"]))
    pool = [product for product in PRODUCTS if product.category == category]
    main_product = weighted_choice([(product, product.weight) for product in pool])
    invoice_products = [main_product]

    if main_product.category in {"Điện thoại", "Tablet"} and random.random() < 0.58:
        accessories = [p for p in PRODUCTS if p.category in {"Phụ kiện", "Tai nghe"}]
        invoice_products.extend(random.sample(accessories, k=random.randint(1, 2)))
    elif main_product.category == "Laptop" and random.random() < 0.42:
        accessories = [p for p in PRODUCTS if p.sub_category in {"Chuột", "Bàn phím", "Củ sạc", "Màn hình"} or p.category == "Màn hình"]
        invoice_products.extend(random.sample(accessories, k=random.randint(1, min(2, len(accessories)))))

    unique = []
    seen = set()
    for product in invoice_products:
        if product.code not in seen:
            unique.append(product)
            seen.add(product.code)
    return unique


def discount_for(customer: Customer | None, product: Product, quantity: int, channel: str) -> int:
    gross = product.list_price * quantity
    rate = 0.0
    if customer and customer.customer_type == "VIP":
        rate += random.uniform(0.03, 0.08)
    elif customer and customer.customer_type == "Thành viên":
        rate += random.uniform(0.01, 0.04)
    elif customer and customer.customer_type == "Khách doanh nghiệp":
        rate += random.uniform(0.04, 0.1)

    if channel in {"Online", "Marketplace"} and random.random() < 0.35:
        rate += random.uniform(0.01, 0.05)
    if product.category == "Phụ kiện" and random.random() < 0.5:
        rate += random.uniform(0.05, 0.18)

    rounded = int(gross * min(rate, 0.22))
    return int(round(rounded / 1000) * 1000)


def generate_invoice_number(invoice_date: date, sequence: int) -> str:
    return f"HD{invoice_date.strftime('%y%m%d')}{sequence:05d}"


def generate_sales_excel(
    output_dir: str,
    customers: list[Customer],
    start_date: date,
    end_date: date,
    days_per_file: int = 1,
) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DanhSachHoaDon"

    headers = [
        "Mã hóa đơn", "Mã sản phẩm", "Mã khách hàng",
        "Mã nhân viên", "Ngày bán", "Số lượng", "Đơn giá",
        "Chiết khấu", "Thuế VAT", "Phương thức TT", "Kênh bán", "Hoàn trả",
    ]
    ws.append(headers)

    invoice_sequence = 1
    row_count = 0
    current = start_date
    daily_units: dict[tuple[date, str], int] = {}

    while current <= end_date:
        is_weekend = current.weekday() >= 5
        is_salary_window = current.day in {1, 2, 3, 25, 26, 27, 28}
        base_invoices = random.randint(18, 34)
        if is_weekend:
            base_invoices += random.randint(8, 18)
        if is_salary_window:
            base_invoices += random.randint(5, 14)

        for _ in range(base_invoices):
            channel = weighted_choice([("POS", 54), ("Online", 28), ("Telesales", 8), ("Marketplace", 10)])
            customer = choose_customer(customers, channel)
            employee = weighted_choice([(emp, 1 if emp.role != "Thu ngân" else 2) for emp in EMPLOYEES])
            invoice_time = realistic_sale_time(current)
            invoice_no = generate_invoice_number(current, invoice_sequence)
            payment = weighted_choice([
                ("Tiền mặt", 27),
                ("Chuyển khoản", 34),
                ("Thẻ", 23),
                ("Ví điện tử", 16),
            ])
            products = choose_products_for_invoice(channel)
            return_invoice = 1 if random.random() < 0.018 else 0

            for product in products:
                if product.category in {"Phụ kiện", "Tai nghe"}:
                    quantity = random.choices([1, 2, 3, 4], weights=[66, 24, 8, 2], k=1)[0]
                else:
                    quantity = random.choices([1, 2, 3], weights=[86, 11, 3], k=1)[0]
                if customer and customer.customer_type == "Khách doanh nghiệp":
                    quantity += random.choices([0, 1, 2, 3], weights=[54, 24, 15, 7], k=1)[0]

                discount = discount_for(customer, product, quantity, channel)
                taxable_amount = max(product.list_price * quantity - discount, 0)
                vat = int(round(taxable_amount * 0.08 / 1000) * 1000)

                ws.append([
                    invoice_no,
                    product.code,
                    customer.code if customer else "",
                    employee.code,
                    format_datetime(invoice_time),
                    -quantity if return_invoice else quantity,
                    product.list_price,
                    -discount if return_invoice else discount,
                    -vat if return_invoice else vat,
                    payment,
                    channel,
                    return_invoice,
                ])
                daily_units[(current, product.code)] = daily_units.get((current, product.code), 0) + (-quantity if return_invoice else quantity)
                row_count += 1

            invoice_sequence += 1

        current += timedelta(days=days_per_file)

    style_sheet(ws)
    for row in ws.iter_rows(min_row=2):
        row[6].number_format = "#,##0"
        row[7].number_format = "#,##0"
        row[8].number_format = "#,##0"

    filename = f"BaoCaoDoanhThu_{start_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"  [OK] {filepath} ({row_count} dòng bán hàng)")
    return row_count


def generate_inventory_excel(
    output_dir: str,
    start_date: date,
    end_date: date,
    days_per_file: int = 1,
) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QuanLyKho"

    headers = [
        "Mã sản phẩm", "Ngày chụp ảnh",
        "Tồn đầu ngày", "Tồn cuối ngày", "Nhập trong ngày",
        "Bán trong ngày", "Điều chỉnh", "Đơn giá vốn",
    ]
    ws.append(headers)

    current_stock = {}
    for product in PRODUCTS:
        if product.category == "Phụ kiện":
            current_stock[product.code] = random.randint(80, 260)
        elif product.category in {"Điện thoại", "Tai nghe"}:
            current_stock[product.code] = random.randint(18, 90)
        else:
            current_stock[product.code] = random.randint(8, 45)

    row_count = 0
    current = start_date
    while current <= end_date:
        for product in PRODUCTS:
            opening = current_stock[product.code]
            reorder_trigger = opening < (45 if product.category == "Phụ kiện" else 12)
            inbound = 0
            if reorder_trigger or random.random() < (0.1 if product.category != "Phụ kiện" else 0.16):
                inbound = random.randint(25, 90) if product.category == "Phụ kiện" else random.randint(5, 28)

            if product.category == "Phụ kiện":
                sold = random.randint(0, 15)
            elif product.category == "Điện thoại":
                sold = random.randint(0, 8)
            else:
                sold = random.randint(0, 4)

            adjustment = random.choices([0, 0, 0, 0, -1, 1, -2], weights=[70, 8, 8, 6, 4, 3, 1], k=1)[0]
            closing = max(0, opening + inbound - sold + adjustment)

            ws.append([
                product.code,
                format_date(current),
                opening,
                closing,
                inbound,
                sold,
                adjustment,
                product.cost,
            ])
            current_stock[product.code] = closing
            row_count += 1

        current += timedelta(days=days_per_file)

    style_sheet(ws)
    for row in ws.iter_rows(min_row=2):
        row[7].number_format = "#,##0"

    filename = f"QuanLyKho_{start_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"  [OK] {filepath} ({row_count} dòng tồn kho)")
    return row_count


def generate_purchase_csv(output_dir: str, start_date: date, end_date: date, num_orders: int = 160) -> list[dict]:
    rows = []
    for index in range(1, num_orders + 1):
        product = weighted_choice([(product, max(1, product.weight)) for product in PRODUCTS])
        order_date = random_date(start_date, end_date)
        lead_days = random.choices([2, 3, 5, 7, 10, 14, 21], weights=[15, 18, 24, 18, 12, 8, 5], k=1)[0]
        received_date = order_date + timedelta(days=lead_days)
        if product.category == "Phụ kiện":
            ordered_qty = random.randint(60, 260)
        elif product.category in {"Điện thoại", "Tai nghe"}:
            ordered_qty = random.randint(12, 70)
        else:
            ordered_qty = random.randint(4, 32)
        shortage = random.choices([0, 0, 0, 1, 2, 3, 5], weights=[50, 20, 12, 8, 5, 3, 2], k=1)[0]
        received_qty = max(0, ordered_qty - shortage)
        unit_cost = int(round(product.cost * random.uniform(0.97, 1.035) / 1000) * 1000)

        rows.append({
            "Số phiếu đặt": f"PN{order_date.strftime('%y%m')}{index:05d}",
            "Mã SP": product.code,
            "Mã NCC": product.supplier_id,
            "Mã cửa hàng": "STORE_DEFAULT",
            "Ngày đặt": format_date(order_date),
            "Ngày nhận": format_date(received_date),
            "Số lượng đặt": ordered_qty,
            "Số lượng nhận": received_qty,
            "Đơn giá nhập": unit_cost,
            "Đã thanh toán": random.choices(["1", "0"], weights=[86, 14], k=1)[0],
        })

    csv_path = os.path.join(output_dir, "PhieuNhapHang.csv")
    write_csv(csv_path, rows)
    print(f"  [OK] {csv_path} ({len(rows)} phiếu nhập)")
    return rows


def main() -> None:
    print("=" * 64)
    print("REALISTIC SAMPLE DATA GENERATOR")
    print("=" * 64)
    print(f"Seed: {RANDOM_SEED}")

    start_date = date(2025, 1, 1)
    end_date = date(2026, 1, 31)
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FOLDER)
    os.makedirs(output_dir, exist_ok=True)

    print(f"\n[output] {output_dir}")
    generate_supplier_csv(output_dir)
    generate_product_csv(output_dir)
    customers = generate_customer_csv(output_dir, num_customers=220)
    sales_rows = generate_sales_excel(output_dir, customers, start_date, end_date)
    inventory_rows = generate_inventory_excel(output_dir, start_date, end_date)
    purchase_rows = generate_purchase_csv(output_dir, start_date, end_date, num_orders=160)

    print(f"\n{'=' * 64}")
    print(f"Tổng dòng bán hàng: {sales_rows:,}")
    print(f"Tổng dòng tồn kho: {inventory_rows:,}")
    print(f"Tổng phiếu nhập: {len(purchase_rows):,}")
    print(f"Thư mục output: {output_dir}")
    print("=" * 64)

    print("\nFile đã tạo:")
    for filename in sorted(os.listdir(output_dir)):
        if filename.startswith("."):
            continue
        filepath = os.path.join(output_dir, filename)
        size_kb = os.path.getsize(filepath) / 1024
        print(f"  {OUTPUT_FOLDER}/{filename} ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
