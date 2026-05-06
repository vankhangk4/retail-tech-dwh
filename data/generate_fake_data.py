#!/usr/bin/env python3
"""
generate_fake_data.py
Sinh fake data cho Data Warehouse Multi-Tenant.
Dựa trên schema SQL: STG_SalesRaw, STG_InventoryRaw, STG_PurchaseRaw,
STG_ProductRaw, STG_CustomerRaw.

Output: Excel/CSV files theo format trong data/fake_data/
"""

import os
import random
from datetime import date, timedelta

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl

# ============================================================
# 1. MASTER DATA
# ============================================================

OUTPUT_FOLDER = "data"

# Danh mục sản phẩm công nghệ (Shared — dùng chung nhiều tenant)
PRODUCTS = [
    # (MaSP, TenSP, ThuongHieu, DanhMuc, DanhMucCon, GiaVon, GiaNiemYet)
    ("SP001", "iPhone 15 Pro 256GB", "Apple", "Điện thoại", "iPhone", 18000000, 25000000),
    ("SP002", "iPhone 15 128GB", "Apple", "Điện thoại", "iPhone", 15000000, 20000000),
    ("SP003", "Samsung Galaxy S24 Ultra", "Samsung", "Điện thoại", "Galaxy S", 17000000, 22000000),
    ("SP004", "Samsung Galaxy A55", "Samsung", "Điện thoại", "Galaxy A", 8000000, 11000000),
    ("SP005", "Xiaomi Redmi Note 13 Pro", "Xiaomi", "Điện thoại", "Redmi Note", 5000000, 7500000),
    ("SP006", "MacBook Pro 14 M3", "Apple", "Laptop", "MacBook Pro", 32000000, 45000000),
    ("SP007", "MacBook Air M2 13-inch", "Apple", "Laptop", "MacBook Air", 20000000, 28000000),
    ("SP008", "Dell XPS 15", "Dell", "Laptop", "XPS", 25000000, 35000000),
    ("SP009", "Lenovo ThinkPad X1 Carbon", "Lenovo", "Laptop", "ThinkPad", 22000000, 32000000),
    ("SP010", "HP Pavilion 15", "HP", "Laptop", "Pavilion", 14000000, 19000000),
    ("SP011", "ASUS ROG Strix G16", "ASUS", "Laptop", "ROG", 28000000, 38000000),
    ("SP012", "iPad Pro 12.9 M4", "Apple", "Tablet", "iPad Pro", 24000000, 32000000),
    ("SP013", "Samsung Galaxy Tab S9 Ultra", "Samsung", "Tablet", "Galaxy Tab", 20000000, 28000000),
    ("SP014", "iPad Air 11 inch M2", "Apple", "Tablet", "iPad Air", 14000000, 19000000),
    ("SP015", "AirPods Pro 2", "Apple", "Tai nghe", "AirPods", 4500000, 6500000),
    ("SP016", "Sony WH-1000XM5", "Sony", "Tai nghe", "Over-ear", 5500000, 8000000),
    ("SP017", "Samsung Galaxy Watch 6", "Samsung", "Đồng hồ", "Galaxy Watch", 6000000, 8500000),
    ("SP018", "Apple Watch Series 9 45mm", "Apple", "Đồng hồ", "Apple Watch", 9000000, 12500000),
    ("SP019", "Apple Watch SE 40mm", "Apple", "Đồng hồ", "Apple Watch", 5000000, 7500000),
    ("SP020", "Cáp sạc Lightning 1m", "Apple", "Phụ kiện", "Cáp sạc", 200000, 350000),
    ("SP021", "Sạc dự phòng 10000mAh", "Anker", "Phụ kiện", "Sạc dự phòng", 350000, 550000),
    ("SP022", "Ốp lưng iPhone 15 Silicone", "Apple", "Phụ kiện", "Ốp lưng", 150000, 299000),
    ("SP023", "Miếng dán kính iPhone 15", "Spigen", "Phụ kiện", "Miếng dán", 80000, 150000),
    ("SP024", "Chuột Logitech MX Master 3S", "Logitech", "Phụ kiện", "Chuột", 1800000, 2800000),
    ("SP025", "Bàn phím Logitech MX Keys", "Logitech", "Phụ kiện", "Bàn phím", 1500000, 2300000),
]

# Nhà cung cấp
SUPPLIERS = [
    ("NCC001", "Apple Việt Nam", "Việt Nam"),
    ("NCC002", "Samsung Electronics Việt Nam", "Việt Nam"),
    ("NCC003", "Công ty TNHH Điện tử Xiaomi Việt Nam", "Việt Nam"),
    ("NCC004", "Dell Technologies Việt Nam", "Việt Nam"),
    ("NCC005", "Lenovo Việt Nam", "Việt Nam"),
    ("NCC006", "HP Việt Nam", "Việt Nam"),
    ("NCC007", "ASUS Việt Nam", "Việt Nam"),
    ("NCC008", "Sony Việt Nam", "Nhật Bản"),
    ("NCC009", "Logitech Việt Nam", "Thụy Sĩ"),
    ("NCC010", "Anker Innovations Việt Nam", "Trung Quốc"),
]

EMPLOYEE_NAMES = [
    "Nguyễn Văn An",
    "Trần Thị Bình",
    "Lê Đức Chung",
    "Phạm Minh Đức",
]

# Khách hàng mẫu (sẽ sinh thêm ngẫu nhiên)
FIRST_NAMES_MALE = ["Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Bùi", "Đặng", "Vũ", "Ngô", "Đỗ", "Trương", "Hồ", "Dương", "Đinh", "Cao"]
FIRST_NAMES_FEMALE = ["Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Bùi", "Đặng", "Vũ", "Ngô", "Đỗ", "Trương", "Hồ", "Dương", "Đinh", "Cao", "Trịnh", "Lương", "Tạ", "Phí", "Chu"]
MIDDLE_NAMES = ["Văn", "Thị", "Đức", "Minh", "Hoàng", "Quang", "Thanh", "Thị", "Thị", "Xuân", "Hồng", "Anh", "Tuấn", "Hải", "LAN"]
LAST_NAMES_MALE = ["An", "Bình", "Chung", "Đức", "Em", "Phượng", "Giỏi", "Hoa", "Hùng", "Dương", "Cường", "Thắng", "Sơn", "Tùng", "Trung", "Hải", "Long", "Nam", "Toàn", "Trí"]
LAST_NAMES_FEMALE = ["Anh", "Bích", "Chi", "Dung", "Hà", "Hương", "Lan", "Linh", "Mai", "My", "Nga", "Oanh", "Phương", "Thảo", "Thúy", "Trang", "Yến", "Hằng", "Ngọc", "Quỳnh"]

CITIES = ["Hà Nội", "Hồ Chí Minh", "Đà Nẵng", "Hải Phòng", "Cần Thơ", "Nha Trang", "Huế", "Bình Dương", "Vũng Tàu", "Quảng Ninh"]
CUSTOMER_TYPES = ["Lẻ", "Sỉ"]
PAYMENT_METHODS = ["Tiền mặt", "Chuyển khoản", "Tín dụng", "Ví điện tử"]
SALES_CHANNELS = ["InStore", "Online", "POS"]
AGE_GROUPS = ["18-24", "25-34", "35-44", "45-54", "55+"]

random.seed(42)


# ============================================================
# 2. HELPER FUNCTIONS
# ============================================================

def random_phone():
    prefixes = ["09", "08", "07", "03", "09"]
    return f"{random.choice(prefixes)}{random.randint(10000000, 99999999)}"


def random_email(first_name, last_name):
    domains = ["gmail.com", "yahoo.com", "outlook.com", "icloud.com"]
    suffix = random.randint(1, 999)
    return f"{first_name.lower()}.{last_name.lower()}{suffix}@{random.choice(domains)}"


def random_date(start: date, end: date):
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, delta))


def format_date(d: date):
    return d.strftime("%d/%m/%Y")


def build_employees():
    return [
        (f"{idx:03d}", full_name)
        for idx, full_name in enumerate(EMPLOYEE_NAMES, start=1)
    ]


def generate_invoice_number(seq: int):
    return f"{seq:06d}"


def generate_order_number(seq: int):
    return f"PN{seq:06d}"


# ============================================================
# 3. GENERATE PRODUCTS (DanhMucSanPham.csv)
# ============================================================

def generate_product_csv(output_dir: str):
    """Sinh file DanhMucSanPham.csv."""
    rows = []
    for i, p in enumerate(PRODUCTS):
        ma_sp, ten_sp, thuong_hieu, danh_muc, danh_muc_con, gia_von, gia_niem_yet = p
        warranty = random.choice([6, 12, 18, 24])
        rows.append({
            "Mã SP": ma_sp,
            "Tên SP": ten_sp,
            "Thương hiệu": thuong_hieu,
            "Danh mục": danh_muc,
            "Danh mục con": danh_muc_con,
            "Giá vốn": gia_von,
            "Giá niêm yết": gia_niem_yet,
            "Đơn vị tính": "cái",
            "Bảo hành (tháng)": warranty,
        })

    # Write CSV
    csv_path = os.path.join(output_dir, "DanhMucSanPham.csv")
    import csv

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    print(f"  [OK] {csv_path} ({len(rows)} sản phẩm)")
    return rows


# ============================================================
# 4. GENERATE CUSTOMERS
# ============================================================

def generate_customer_csv(output_dir: str, num_customers: int = 50):
    """Sinh file DanhMucKhachHang.csv."""
    rows = []
    base_date = date(2024, 1, 1)
    end_date = date(2026, 1, 1)

    for i in range(num_customers):
        gender = random.choice(["M", "F"])
        first = random.choice(FIRST_NAMES_MALE if gender == "M" else FIRST_NAMES_FEMALE)
        last = random.choice(LAST_NAMES_MALE if gender == "M" else LAST_NAMES_FEMALE)
        mid = random.choice(MIDDLE_NAMES)
        full_name = f"{first} {mid} {last}"
        ma_kh = f"{i+1:05d}"

        dob_year = random.randint(1980, 2006)
        dob_month = random.randint(1, 12)
        dob_day = random.randint(1, 28)
        dob = f"{dob_day:02d}/{dob_month:02d}/{dob_year}"

        age = 2026 - dob_year
        if age < 25:
            age_group = "18-24"
        elif age < 35:
            age_group = "25-34"
        elif age < 45:
            age_group = "35-44"
        elif age < 55:
            age_group = "45-54"
        else:
            age_group = "55+"

        city = random.choice(CITIES)
        province = city
        member_since = random_date(base_date, end_date)
        loyalty_points = random.randint(0, 5000000)
        customer_type = random.choice(CUSTOMER_TYPES)

        rows.append({
            "Mã KH": ma_kh,
            "Họ tên": full_name,
            "Giới tính": gender,
            "Ngày sinh": dob,
            "Thành phố": city,
            "Tỉnh/TP": province,
            "Số điện thoại": random_phone(),
            "Email": random_email(first, last),
            "Loại KH": customer_type,
            "Điểm tích lũy": loyalty_points,
            "Ngày đăng ký": format_date(member_since),
            "Nhóm tuổi": age_group,
        })

    csv_path = os.path.join(output_dir, "DanhMucKhachHang.csv")
    import csv

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    print(f"  [OK] {csv_path} ({len(rows)} khách hàng)")
    return rows


# ============================================================
# 5. GENERATE SALES
# ============================================================

def generate_sales_excel(
    output_dir: str,
    start_date: date,
    end_date: date,
    days_per_file: int = 1,
):
    """
    Sinh file BaoCaoDoanhThu_YYYYMMDD.xlsx
    Sheet: "DanhSachHoaDon"
    Mỗi file chứa dữ liệu N ngày.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DanhSachHoaDon"

    # Header
    headers = [
        "Mã hóa đơn", "Mã sản phẩm", "Mã khách hàng",
        "Mã nhân viên", "Ngày bán", "Số lượng", "Đơn giá",
        "Chiết khấu", "Thuế VAT", "Phương thức TT", "Kênh bán", "Hoàn trả",
    ]
    ws.append(headers)

    employees = build_employees()
    num_employees = len(employees)
    num_products = len(PRODUCTS)

    invoice_counter = 1
    row_count = 0

    current = start_date
    while current <= end_date:
        # Mỗi ngày: 5-20 hóa đơn, mỗi hóa đơn: 1-5 dòng sản phẩm
        num_invoices = random.randint(8, 25)

        for _ in range(num_invoices):
            # Chọn ngày trong khoảng (để tạo vài ngày nghỉ)
            invoice_date = current - timedelta(days=random.randint(0, 0))
            ma_hd = generate_invoice_number(invoice_counter)
            ma_nv = employees[random.randint(0, num_employees - 1)][0]

            # 1-4 dòng sản phẩm trong hóa đơn
            num_lines = random.randint(1, 4)
            selected_products = random.sample(range(num_products), min(num_lines, num_products))

            for pi in selected_products:
                sp = PRODUCTS[pi]
                ma_sp = sp[0]
                don_gia_ban = sp[6]

                # Khách hàng ngẫu nhiên (hoặc null ~20%)
                if random.random() > 0.15:
                    kh_seq = random.randint(1, 50)
                    ma_kh = f"{kh_seq:05d}"
                else:
                    ma_kh = ""

                so_luong = random.randint(1, 3)

                # Chiết khấu: ~60% có chiết khấu
                if random.random() < 0.6:
                    chiet_khau = int(don_gia_ban * so_luong * random.uniform(0.03, 0.15))
                else:
                    chiet_khau = 0

                # Thuế VAT: 0 hoặc 10% của (doanh thu - chiết khấu)
                if random.random() < 0.3:
                    thue_vat = int((don_gia_ban * so_luong - chiet_khau) * 0.1)
                else:
                    thue_vat = 0

                phuong_thuc_tt = random.choice(PAYMENT_METHODS)
                kenh_ban = random.choice(SALES_CHANNELS)

                # Hoàn trả: ~5% là hoàn trả
                is_hoan_tra = 1 if random.random() < 0.05 else 0

                row = [
                    ma_hd,
                    ma_sp,
                    ma_kh,
                    ma_nv,
                    format_date(invoice_date),
                    so_luong,
                    don_gia_ban,
                    chiet_khau,
                    thue_vat,
                    phuong_thuc_tt,
                    kenh_ban,
                    is_hoan_tra,
                ]
                ws.append(row)
                row_count += 1

            invoice_counter += 1

        current += timedelta(days=days_per_file)

    # Format header
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)

    # Format number columns
    for row in ws.iter_rows(min_row=2):
        row[6].number_format = "#,##0"  # Đơn giá
        row[7].number_format = "#,##0"  # Chiết khấu
        row[8].number_format = "#,##0"  # Thuế VAT

    filename = f"BaoCaoDoanhThu_{start_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"  [OK] {filepath} ({row_count} dòng)")
    return row_count


# ============================================================
# 6. GENERATE INVENTORY
# ============================================================

def generate_inventory_excel(
    output_dir: str,
    start_date: date,
    end_date: date,
    days_per_file: int = 7,
):
    """
    Sinh file QuanLyKho_YYYYMMDD.xlsx
    Mỗi sản phẩm có 1 dòng mỗi ngày ghi nhận tồn kho.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QuanLyKho"

    headers = [
        "Mã sản phẩm", "Ngày chụp ảnh",
        "Tồn đầu ngày", "Tồn cuối ngày", "Nhập trong ngày",
        "Bán trong ngày", "Điều chỉnh", "Đơn giá vốn",
    ]
    ws.append(headers)

    row_count = 0
    current = start_date

    # Tồn kho ban đầu ngẫu nhiên cho mỗi sản phẩm
    current_stock = {p[0]: random.randint(5, 50) for p in PRODUCTS}

    while current <= end_date:
        for sp in PRODUCTS:
            ma_sp = sp[0]
            gia_von = sp[5]

            # Tồn đầu ngày
            opening = current_stock[ma_sp]

            # Nhập trong ngày: random, thỉnh thoảng có nhập
            nhap = random.randint(0, 15) if random.random() < 0.2 else 0

            # Bán trong ngày
            ban = random.randint(0, 8)

            # Điều chỉnh: thỉnh thoảng có mất mát/hỏng
            dieu_chinh = random.randint(-3, 3) if random.random() < 0.1 else 0

            # Tồn cuối ngày
            closing = max(0, opening + nhap - ban + dieu_chinh)

            row = [
                ma_sp,
                format_date(current),
                opening,
                closing,
                nhap,
                ban,
                dieu_chinh,
                gia_von,
            ]
            ws.append(row)
            row_count += 1

            # Cập nhật tồn cho ngày tiếp theo
            current_stock[ma_sp] = closing

        current += timedelta(days=days_per_file)

    # Format header
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)

    # Format number columns
    for row in ws.iter_rows(min_row=2):
        row[7].number_format = "#,##0"  # Đơn giá vốn

    filename = f"QuanLyKho_{start_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"  [OK] {filepath} ({row_count} dòng)")
    return row_count


# ============================================================
# 7. GENERATE PURCHASE
# ============================================================

def generate_purchase_csv(output_dir: str, num_orders: int = 30):
    """
    Sinh file PhieuNhapHang.csv.
    """
    rows = []
    start_date = date(2025, 1, 1)
    end_date = date(2026, 1, 31)

    for i in range(num_orders):
        ma_sp = random.choice(PRODUCTS)[0]
        ma_ncc = random.choice(SUPPLIERS)[0]
        ngay_dat = random_date(start_date, end_date)
        # Lead time: 1-30 ngày
        lead = random.randint(1, 30)
        ngay_nhan = ngay_dat + timedelta(days=lead)
        so_luong_dat = random.randint(10, 100)
        # Thường nhận đủ hoặc thiếu 1 ít
        so_luong_nhan = so_luong_dat - random.randint(0, 3)
        don_gia_nhap = random.randint(500000, 35000000)

        rows.append({
            "Số phiếu đặt": generate_order_number(i + 1),
            "Mã SP": ma_sp,
            "Mã NCC": ma_ncc,
            "Ngày đặt": format_date(ngay_dat),
            "Ngày nhận": format_date(ngay_nhan),
            "Số lượng đặt": so_luong_dat,
            "Số lượng nhận": so_luong_nhan,
            "Đơn giá nhập": don_gia_nhap,
        })

    csv_path = os.path.join(output_dir, "PhieuNhapHang.csv")
    import csv

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    print(f"  [OK] {csv_path} ({len(rows)} phiếu nhập)")
    return rows


# ============================================================
# 8. GENERATE SUPPLIERS CSV (dùng chung)
# ============================================================

def generate_supplier_csv(output_dir: str):
    """Sinh file DanhMucNhaCungCap.csv"""
    rows = []
    for ma, ten, nuoc in SUPPLIERS:
        rows.append({
            "Mã NCC": ma,
            "Tên NCC": ten,
            "Quốc gia": nuoc,
            "Điều khoản TT (ngày)": random.choice([15, 30, 45, 60]),
            "Số điện thoại": random_phone(),
            "Email": f"contact.{ma.lower()}@company.com",
        })

    csv_path = os.path.join(output_dir, "DanhMucNhaCungCap.csv")
    import csv

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    print(f"  [OK] {csv_path} ({len(rows)} nhà cung cấp)")
    return rows


# ============================================================
# 9. MAIN
# ============================================================

def main():
    print("=" * 60)
    print("FAKE DATA GENERATOR")
    print("=" * 60)

    # Khoảng thời gian dữ liệu
    start_date = date(2025, 1, 1)
    end_date = date(2026, 1, 31)

    total_rows = 0
    output_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), OUTPUT_FOLDER
    )
    os.makedirs(output_dir, exist_ok=True)

    print(f"\n[output] {OUTPUT_FOLDER}")

    # 1. Products
    generate_product_csv(output_dir)

    # 2. Customers
    generate_customer_csv(output_dir, num_customers=50)

    # 3. Sales
    rows = generate_sales_excel(output_dir, start_date, end_date, days_per_file=1)
    total_rows += rows

    # 4. Inventory
    generate_inventory_excel(output_dir, start_date, end_date, days_per_file=7)

    # 5. Purchase orders
    generate_purchase_csv(output_dir, num_orders=30)

    # 6. Suppliers
    generate_supplier_csv(output_dir)

    print(f"\n{'=' * 60}")
    print(f"TỔNG CỘNG: {total_rows} dòng hóa đơn được sinh")
    print(f"Thư mục output: data/{OUTPUT_FOLDER}/")
    print("=" * 60)

    # Danh sách file đã tạo
    print("\nFile đã tạo:")
    for f in sorted(os.listdir(output_dir)):
        if f.startswith("."):
            continue
        fpath = os.path.join(output_dir, f)
        size = os.path.getsize(fpath)
        print(f"  {OUTPUT_FOLDER}/{f}  ({size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
